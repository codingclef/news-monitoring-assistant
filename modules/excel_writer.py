from __future__ import annotations

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 일람 시트 전용 컬럼 (분류이유 없음, 분류결과 있음)
COLUMNS_ILAM  = ["No.", "키워드", "날짜/시간", "언론사", "기사제목", "링크", "분류결과"]
COL_WIDTHS_ILAM = [6, 15, 20, 15, 55, 45, 20]

# 그 외 시트 컬럼 (분류이유 있음)
COLUMNS_OTHER = ["No.", "키워드", "날짜/시간", "언론사", "기사제목", "링크", "분류이유"]
COL_WIDTHS_OTHER = [6, 15, 20, 15, 55, 45, 45]

HEADER_BG = "4472C4"
HEADER_FG = "FFFFFF"

SHEET_ILAM = "일람"
SHEET_UNCLASSIFIED = "보류"


def create_excel(
    articles: list[dict],
    category_names: list[str],
) -> bytes:
    """
    시트 순서:
    1. 일람      - 모든 기사 (No. / 분류결과 컬럼 포함)
    2. 사용자 정의 카테고리
    3. 보류      - 분류 불가 기사
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1. 일람 시트
    ws_ilam = wb.create_sheet(title=SHEET_ILAM)
    _setup_header(ws_ilam, is_ilam=True)
    for idx, article in enumerate(articles, 1):
        _add_row(ws_ilam, article, idx, is_ilam=True)
    _apply_column_widths(ws_ilam, is_ilam=True)

    # 2. 사용자 정의 카테고리 시트
    for cat_name in category_names:
        ws = wb.create_sheet(title=cat_name[:31])
        _setup_header(ws, is_ilam=False)
        idx = 1
        for article in articles:
            if article.get("category") == cat_name:
                _add_row(ws, article, idx, is_ilam=False)
                idx += 1
        _apply_column_widths(ws, is_ilam=False)

    # 3. 보류 시트
    ws_unc = wb.create_sheet(title=SHEET_UNCLASSIFIED)
    _setup_header(ws_unc, is_ilam=False)
    idx = 1
    for article in articles:
        if article.get("category") == SHEET_UNCLASSIFIED:
            _add_row(ws_unc, article, idx, is_ilam=False)
            idx += 1
    _apply_column_widths(ws_unc, is_ilam=False)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def _setup_header(ws, is_ilam: bool) -> None:
    columns = COLUMNS_ILAM if is_ilam else COLUMNS_OTHER
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_font = Font(bold=True, color=HEADER_FG, size=11)

    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _add_row(ws, article: dict, row_num: int, is_ilam: bool) -> None:
    row = ws.max_row + 1

    pub_dt = article.get("published_at")
    date_str = pub_dt.strftime("%Y-%m-%d %H:%M") if isinstance(pub_dt, datetime) else ""

    last_col_value = article.get("category", "") if is_ilam else article.get("reason", "")

    values = [
        row_num,
        article.get("keyword", ""),
        date_str,
        article.get("source", ""),
        article.get("title", ""),
        article.get("link", ""),
        last_col_value,
    ]

    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        if col == 6 and value:  # 링크 컬럼
            try:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")
            except Exception:
                pass

    ws.row_dimensions[row].height = 40


def _apply_column_widths(ws, is_ilam: bool) -> None:
    widths = COL_WIDTHS_ILAM if is_ilam else COL_WIDTHS_OTHER
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
